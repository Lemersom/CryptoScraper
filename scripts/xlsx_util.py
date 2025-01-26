from openpyxl import Workbook, load_workbook

def read_xlsx(input_file: str) -> list[dict]:
    """
    Reads an XLSX file and returns a list of dictionaries where each dictionary
    represents a row with column headers as keys.

    Args:
        input_file (str): Path to the input XLSX file.
    
    Returns:
        list[dict]: A list of dictionaries containing the data from the XLSX file.
                    Returns an empty list if an error occurs.
    """
    input_data = []
    try:
        workbook = load_workbook(input_file)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]] # Extract headers from the first row

        # Iterate over the remaining rows and create dictionaries
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            input_data.append(row_dict)
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
    except Exception as e:
        print(f"An error occurred while reading the xlsx file: {e}")
    
    return input_data


def write_xlsx(output_file: str, data: list[dict]) -> None:
    """
    Writes a list of dictionaries to an XLSX file. The keys of the dictionaries
    are used as column headers.

    Args:
        output_file (str): Path to the output XLSX file.
        data (list[dict]): A list of dictionaries to write, where keys are used as headers.
    """
    if not data:
        print("Error: No data to write to the XLSX file.")
        return

    try:
        workbook = Workbook()
        sheet = workbook.active

        headers = list(data[0].keys())
        sheet.append(headers)

        for row in data:
            sheet.append(list(row.values()))
        
        workbook.save(output_file)
    except Exception as e:
        print(f"An error occurred while writing to the xlsx file: {e}")